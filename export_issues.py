# -*- coding: utf-8 -*-
"""
issues 表数据导出脚本
用法：python3 export_issues.py
依赖：openpyxl（pip install openpyxl）
输出：exports/issues_export_YYYYMMDD.xlsx
"""

import sqlite3
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.path.dirname(__file__), 'backend', 'cdrl.db')
EXPORT_DIR = os.path.join(os.path.dirname(__file__), 'exports')

# 列定义：(表头, 取值函数)
COLUMNS = [
    ('检查时间',        lambda r: r['inspection_date'] or r['check_date']),
    ('检查单位',        lambda r: r['inspection_unit'] or r['check_unit']),
    ('建设单位',        lambda r: r['builder_unit']),
    ('检查项目',        lambda r: r['project_name']),
    ('标段',           lambda r: r['section_name']),
    ('施工单位',        lambda r: r['contractor'] or r['contractor_unit'] or ''),
    ('监理单位',        lambda r: r['supervisor'] or r['supervisor_unit'] or ''),
    ('工点名称',        lambda r: r['site_name']),
    ('问题描述',        lambda r: r['description']),
    ('一级分类',        lambda r: r['issue_category']),
    ('二级分类',        lambda r: r['issue_type_level1']),
    ('三级分类',        lambda r: r['issue_type_level2']),
    ('严重程度',        lambda r: r['severity']),
    ('整改期限',        lambda r: r['rectification_deadline']),
    ('是否下发整改通知单', lambda r: '是' if r['is_rectification_notice'] else '否'),
    ('责任单位',        lambda r: r['responsible_unit']),
    ('信用评价扣分(分）', lambda r: r['credit_score']),
    ('检查通知书',      lambda r: r['notice_number']),
    ('检查人',         lambda r: r['check_personnel']),
]

QUERY = """
    SELECT
        i.section_name, i.site_name, i.contractor, i.supervisor,
        i.issue_category, i.issue_type_level1, i.issue_type_level2,
        i.description, i.severity, i.keywords,
        i.inspection_unit, i.inspection_date,
        i.rectification_requirements, i.rectification_deadline,
        i.is_rectification_notice, i.is_bad_behavior_notice,
        i.responsible_unit, i.credit_score,
        s.contractor_unit, s.supervisor_unit,
        p.project_name, p.builder_unit,
        n.notice_number, n.check_date, n.check_unit, n.check_personnel
    FROM issues i
    LEFT JOIN sections s ON i.section_id = s.id
    LEFT JOIN projects p ON s.project_id = p.id
    LEFT JOIN supervision_notices n ON i.supervision_notice_id = n.id
    ORDER BY i.inspection_date, i.section_name, i.site_name
"""


def style_header(ws, num_cols):
    header_fill = PatternFill('solid', fgColor='366092')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='FFFFFF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border


def style_data(ws, num_rows, num_cols):
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_even = PatternFill('solid', fgColor='EBF3FB')
    font = Font(size=10)
    for row in range(2, num_rows + 2):
        fill = fill_even if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=(col == 9))  # 问题描述列自动换行
            if fill:
                cell.fill = fill


def set_col_widths(ws):
    widths = [12, 12, 18, 16, 10, 14, 14, 20, 50, 10, 12, 12, 8, 12, 12, 14, 10, 22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30


def main():
    os.makedirs(EXPORT_DIR, exist_ok=True)
    filename = 'issues_export_%s.xlsx' % datetime.now().strftime('%Y%m%d')
    output_path = os.path.join(EXPORT_DIR, filename)

    print('连接数据库：%s' % DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    print('查询数据...')
    rows = conn.execute(QUERY).fetchall()
    conn.close()
    print('共 %d 条记录' % len(rows))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '问题记录'
    ws.freeze_panes = 'A2'

    # 写表头
    headers = [col[0] for col in COLUMNS]
    ws.append(headers)

    # 写数据
    for row in rows:
        r = dict(row)
        data = []
        for _, fn in COLUMNS:
            try:
                val = fn(r)
            except Exception:
                val = None
            data.append(val if val is not None else '')
        ws.append(data)

    # 样式
    style_header(ws, len(COLUMNS))
    style_data(ws, len(rows), len(COLUMNS))
    set_col_widths(ws)

    wb.save(output_path)
    print('导出完成：%s' % output_path)


if __name__ == '__main__':
    main()

